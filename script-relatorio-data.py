#!/usr/bin/env python3
"""
Gera um relatorio consolidado em texto a partir de planilhas XLSX exportadas do
Twilio (ou formato equivalente). O script agrupa mensagens semelhantes e imprime
no terminal um sumario com todas as mensagens e estatisticas por data/status.

Resumo do fluxo:
1. Le as planilhas informadas (primeira aba por padrao, pode ser alterada).
2. Normaliza e agrupa mensagens cujo corpo seja 70% semelhante (configuravel).
3. Converte o status original para:
      - delivered  -> Entregue
      - qualquer outro -> Enviada e Nao Entregue
4. Consolida contagens por data e status para cada agrupamento.
5. Imprime um relatorio textual completo com overview e todos os agrupamentos.

Dependencias:
    python >= 3.9
    openpyxl (leitura de XLSX)

Uso esperado:
    python3 script-relatorio-data.py relatorio-*.xlsx --threshold 0.7
"""

from __future__ import annotations

import argparse
import datetime as dt
import glob
import os
import re
import sys
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from difflib import SequenceMatcher
import textwrap
from typing import Dict, List, Optional, Sequence

try:
    from openpyxl import load_workbook  # type: ignore
except ImportError as exc:  # pragma: no cover - tratamos erros em tempo de execucao
    sys.stderr.write(
        "[ERRO] Dependencia ausente: openpyxl. Instale com 'pip install openpyxl'.\n"
    )
    raise

# --------------------------------------------------------------------------------------
# Modelos de dados
# --------------------------------------------------------------------------------------

NORMALIZED_STATUS_DELIVERED = "Entregue"
NORMALIZED_STATUS_OTHER = "Enviada e Não Entregue"

URL_PATTERN = re.compile(r"https?://\S+", re.IGNORECASE)
URL_PATTERN_ALT = re.compile(r"\bhttps?//\S+", re.IGNORECASE)
NON_WORD_PATTERN = re.compile(r"[^\w\s]")
MULTI_WS_PATTERN = re.compile(r"\s+")


@dataclass
class Message:
    """Representa a linha relevante do relatorio."""

    source_file: str
    body: str
    body_normalized: str
    status_raw: str
    status_mapped: str
    date: Optional[dt.date]
    metadata: Dict[str, object] = field(default_factory=dict)


@dataclass
class MessageCluster:
    """Agrupamento de mensagens por similaridade de texto."""

    representative: str
    representative_raw: str
    messages: List[Message] = field(default_factory=list)

    def add(self, message: Message) -> None:
        self.messages.append(message)

    @property
    def total(self) -> int:
        return len(self.messages)

    def counts_by_date(self) -> Dict[dt.date, Dict[str, int]]:
        counts: Dict[dt.date, Dict[str, int]] = defaultdict(lambda: defaultdict(int))
        for msg in self.messages:
            if msg.date is None:
                continue
            counts[msg.date][msg.status_mapped] += 1
        return counts

    def counts_totals(self) -> Dict[str, int]:
        totals: Dict[str, int] = defaultdict(int)
        for msg in self.messages:
            totals[msg.status_mapped] += 1
        return totals

    def body_template(self) -> str:
        """Cria um texto exemplo substituindo variacoes por {variavel}."""
        if not self.messages:
            return ""

        tokens_per_message = [msg.body.strip().split() for msg in self.messages if msg.body.strip()]
        if not tokens_per_message:
            return ""

        if len(tokens_per_message) == 1:
            return " ".join(tokens_per_message[0])

        max_len = max(len(tokens) for tokens in tokens_per_message)
        template_tokens: List[str] = []

        for idx in range(max_len):
            tokens_at_idx = [
                tokens[idx] if idx < len(tokens) else None for tokens in tokens_per_message
            ]
            has_missing = any(token is None for token in tokens_at_idx)
            values = {token for token in tokens_at_idx if token is not None}

            if not values and has_missing:
                break

            if len(values) == 1 and not has_missing:
                template_tokens.append(next(iter(values)))
            else:
                if not template_tokens or template_tokens[-1] != "{variavel}":
                    template_tokens.append("{variavel}")

        return " ".join(template_tokens)


# --------------------------------------------------------------------------------------
# Utilidades de leitura e transformacao
# --------------------------------------------------------------------------------------

def discover_input_files(patterns: Sequence[str]) -> List[str]:
    """Expande padroes/globs e devolve apenas arquivos existentes."""
    found: List[str] = []
    for item in patterns:
        if any(char in item for char in "*?[]"):
            expanded = glob.glob(item)
            found.extend(expanded)
        else:
            if os.path.exists(item):
                found.append(item)
            else:
                sys.stderr.write(f"[AVISO] Arquivo nao encontrado: {item}\n")
    unique_sorted = sorted(set(found))
    if not unique_sorted:
        raise FileNotFoundError("Nenhum arquivo XLSX encontrado a partir dos padroes informados.")
    return unique_sorted


def normalize_header(value: object) -> str:
    text = str(value or "").strip().lower()
    return re.sub(r"\s+", "_", text)


def normalize_body(text: object) -> str:
    raw = str(text or "").strip()
    if not raw:
        return ""

    # Remove URLs e tokens que variam muito para reduzir variancia.
    temp = URL_PATTERN.sub(" ", raw)
    temp = URL_PATTERN_ALT.sub(" ", temp)
    temp = NON_WORD_PATTERN.sub(" ", temp)
    temp = MULTI_WS_PATTERN.sub(" ", temp).strip().lower()
    return temp


def map_status(value: object) -> str:
    status = str(value or "").strip().lower()
    if status == "delivered":
        return NORMALIZED_STATUS_DELIVERED
    return NORMALIZED_STATUS_OTHER


def excel_serial_to_datetime(serial: float) -> dt.datetime:
    """Converte numero serial do Excel para datetime."""
    base = dt.datetime(1899, 12, 30)
    return base + dt.timedelta(days=serial)


def parse_datetime(value: object) -> Optional[dt.datetime]:
    """Recebe objetos variados e devolve datetime ou None."""
    if value is None:
        return None
    if isinstance(value, dt.datetime):
        return value
    if isinstance(value, dt.date):
        return dt.datetime.combine(value, dt.time.min)
    if isinstance(value, (int, float)):
        try:
            return excel_serial_to_datetime(float(value))
        except Exception:
            return None
    text = str(value).strip()
    if not text:
        return None

    # Tenta formato ISO 8601 (com ou sem timezone)
    try:
        # fromisoformat lida com "YYYY-MM-DDTHH:MM:SS+-HH:MM"
        return dt.datetime.fromisoformat(text)
    except ValueError:
        pass

    # Outras tentativas comuns
    for fmt in ("%Y-%m-%d %H:%M:%S", "%d/%m/%Y %H:%M:%S", "%Y-%m-%d"):
        try:
            return dt.datetime.strptime(text, fmt)
        except ValueError:
            continue

    return None


def extract_date(row: Dict[str, object]) -> Optional[dt.date]:
    """Extrai a data mais confiavel (date_sent > date_created)."""
    for key in ("date_sent", "data_envio", "sent_at", "date_created", "data_criacao"):
        if key not in row:
            continue
        dt_obj = parse_datetime(row[key])
        if dt_obj is not None:
            return dt_obj.date()
    return None


def load_messages_from_workbook(path: str, sheet_name: Optional[str] = None) -> List[Message]:
    """Le uma planilha XLSX e devolve a lista de mensagens validas."""
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[sheet_name] if sheet_name else workbook.active

    rows_iter = worksheet.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return []

    header = [normalize_header(cell) for cell in header_row]
    messages: List[Message] = []

    for raw_row in rows_iter:
        if not any(raw_row):
            continue  # linha vazia

        row_dict = dict(zip(header, raw_row))

        body_raw = row_dict.get("body") or row_dict.get("texto")
        if not body_raw:
            continue

        status_raw = row_dict.get("status", "")
        status_mapped = map_status(status_raw)
        date_value = extract_date(row_dict)

        metadata = {
            "to": row_dict.get("to") or row_dict.get("destino"),
            "status_original": status_raw,
            "date_sent": row_dict.get("date_sent"),
            "date_created": row_dict.get("date_created"),
            "sid": row_dict.get("sid"),
        }

        message = Message(
            source_file=os.path.basename(path),
            body=str(body_raw).strip(),
            body_normalized=normalize_body(body_raw),
            status_raw=str(status_raw or ""),
            status_mapped=status_mapped,
            date=date_value,
            metadata=metadata,
        )
        messages.append(message)

    return messages


def load_messages(paths: Sequence[str], sheet_name: Optional[str] = None) -> List[Message]:
    all_messages: List[Message] = []
    for path in paths:
        try:
            msgs = load_messages_from_workbook(path, sheet_name=sheet_name)
        except KeyError:
            sys.stderr.write(
                f"[AVISO] Aba '{sheet_name}' nao encontrada em {path}. Utilizando a primeira aba disponivel.\n"
            )
            msgs = load_messages_from_workbook(path, sheet_name=None)
        except Exception as exc:
            sys.stderr.write(f"[ERRO] Falha ao ler {path}: {exc}\n")
            continue
        all_messages.extend(msgs)
    return all_messages


# --------------------------------------------------------------------------------------
# Agrupamento por similaridade
# --------------------------------------------------------------------------------------

def are_similar(a: str, b: str, threshold: float) -> bool:
    if not a or not b:
        return False
    matcher = SequenceMatcher(None, a, b, autojunk=False)
    if matcher.real_quick_ratio() < threshold:
        return False
    if matcher.quick_ratio() < threshold:
        return False
    return matcher.ratio() >= threshold


def cluster_messages(messages: Sequence[Message], threshold: float) -> List[MessageCluster]:
    clusters: List[MessageCluster] = []
    exact_index: Dict[str, MessageCluster] = {}
    prefix_index: Dict[str, List[MessageCluster]] = defaultdict(list)

    for msg in messages:
        normalized = msg.body_normalized
        prefix = normalized[:40] if normalized else ""

        assigned_cluster: Optional[MessageCluster] = None

        # Caminho rapido: match exato ja conhecido.
        if normalized and normalized in exact_index:
            assigned_cluster = exact_index[normalized]

        if assigned_cluster is None:
            candidates = prefix_index.get(prefix, [])
            evaluated_ids = {id(cluster) for cluster in candidates}

            for cluster in candidates:
                if are_similar(normalized, cluster.representative, threshold):
                    assigned_cluster = cluster
                    break

            if assigned_cluster is None:
                for cluster in clusters:
                    if id(cluster) in evaluated_ids:
                        continue
                    rep = cluster.representative
                    if normalized and rep:
                        max_len = max(len(rep), len(normalized))
                        if max_len and min(len(rep), len(normalized)) / max_len < threshold - 0.15:
                            continue  # pulo heuristico
                    if are_similar(normalized, rep, threshold):
                        assigned_cluster = cluster
                        break

        if assigned_cluster is None:
            assigned_cluster = MessageCluster(
                representative=normalized,
                representative_raw=msg.body,
                messages=[],
            )
            clusters.append(assigned_cluster)

        assigned_cluster.add(msg)

        if normalized and normalized not in exact_index:
            exact_index[normalized] = assigned_cluster

        bucket = prefix_index[prefix]
        if assigned_cluster not in bucket:
            bucket.append(assigned_cluster)

    # Ordena clusters por volume (maior primeiro)
    clusters.sort(key=lambda c: c.total, reverse=True)
    return clusters


# --------------------------------------------------------------------------------------
# Relatorio textual
# --------------------------------------------------------------------------------------

def format_int(value: int) -> str:
    """Formata inteiros usando ponto como separador de milhar."""
    return f"{value:,}".replace(",", ".")


def summarize_overall(clusters: Sequence[MessageCluster]) -> Dict[str, object]:
    total_msgs = sum(cluster.total for cluster in clusters)
    total_clusters = len(clusters)
    date_counts: Dict[dt.date, int] = defaultdict(int)
    delivered = 0
    sent_not_delivered = 0

    for cluster in clusters:
        for msg in cluster.messages:
            if msg.date:
                date_counts[msg.date] += 1
            if msg.status_mapped == NORMALIZED_STATUS_DELIVERED:
                delivered += 1
            else:
                sent_not_delivered += 1

    return {
        "total_messages": total_msgs,
        "total_clusters": total_clusters,
        "total_entregue": delivered,
        "total_nao_entregue": sent_not_delivered,
        "dates": dict(date_counts),
    }


def sanitize_template(cluster: MessageCluster) -> str:
    """
    Ajusta o template para evitar que todo o texto seja reduzido a {variavel}.
    Quando o texto dinamico e muito grande, utiliza a mensagem representativa.
    """
    template = cluster.body_template().strip()
    if not template:
        return cluster.representative_raw

    tokens = template.split()
    if not tokens:
        return cluster.representative_raw

    var_tokens = sum(1 for token in tokens if token == "{variavel}")
    if var_tokens == len(tokens):
        return cluster.representative_raw

    if var_tokens / len(tokens) >= 0.6:
        return cluster.representative_raw

    return template


def collect_sample_messages(cluster: MessageCluster, limit: int = 3) -> List[str]:
    """Seleciona as mensagens mais frequentes do agrupamento (limit padrao = 3)."""
    counter = Counter(msg.body.strip() for msg in cluster.messages if msg.body.strip())
    samples = [text for text, _ in counter.most_common(limit)]
    return samples


def format_message_block(text: str, width: int = 96, indent: str = "    ") -> str:
    """Quebra e indenta mensagens longas para leitura mais facil no terminal."""
    clean = MULTI_WS_PATTERN.sub(" ", text.strip())
    wrapped = textwrap.fill(clean, width=width)
    return textwrap.indent(wrapped, indent)


def print_overall_summary(overview: Dict[str, object], files: Sequence[str], threshold: float) -> None:
    print("=" * 100)
    print(f"Relatorio consolidado - similaridade >= {int(threshold * 100)}%")
    print("=" * 100)
    print(f"Mensagens totais : {format_int(int(overview['total_messages']))}")
    print(f"Agrupamentos     : {format_int(int(overview['total_clusters']))}")
    print(f"Entregues        : {format_int(int(overview['total_entregue']))}")
    print(f"Nao entregues    : {format_int(int(overview['total_nao_entregue']))}")

    date_counts: Dict[dt.date, int] = overview["dates"]  # type: ignore[assignment]
    if date_counts:
        print("\nDatas com mensagens:")
        for date, count in sorted(date_counts.items()):
            print(f"  - {date.isoformat()}: {format_int(int(count))}")
    else:
        print("\nDatas com mensagens: nao informadas.")

    if files:
        print("\nArquivos analisados:")
        for name in files:
            print(f"  - {name}")
    print()


def print_cluster_details(clusters: Sequence[MessageCluster]) -> None:
    print("Agrupamentos detalhados")
    print("-" * 100)
    for idx, cluster in enumerate(clusters, start=1):
        totals = cluster.counts_totals()
        delivered = totals.get(NORMALIZED_STATUS_DELIVERED, 0)
        other = sum(totals.values()) - delivered

        print(
            f"[{idx}] Total: {format_int(cluster.total)} | Entregues: {format_int(delivered)} | Nao Entregues: {format_int(other)}"
        )

        template = sanitize_template(cluster)
        print("  Texto base:")
        print(format_message_block(template, indent="    "))

        samples = collect_sample_messages(cluster)
        if samples:
            print("  Exemplos de mensagens:")
            for sample in samples:
                print(format_message_block(sample, indent="    - "))

        date_counts = cluster.counts_by_date()
        if date_counts:
            print("  Datas:")
            for date, status_counts in sorted(date_counts.items()):
                total = sum(status_counts.values())
                delivered_date = status_counts.get(NORMALIZED_STATUS_DELIVERED, 0)
                other_date = total - delivered_date
                print(
                    f"    - {date.isoformat()}: {format_int(total)} (Entregues: {format_int(delivered_date)}, Nao Entregues: {format_int(other_date)})"
                )
        else:
            print("  Datas: nao informadas.")

        source_files = sorted({msg.source_file for msg in cluster.messages})
        if source_files:
            print("  Arquivos de origem:")
            for name in source_files:
                print(f"    - {name}")

        print("-" * 100)


def print_console_report(
    clusters: Sequence[MessageCluster],
    threshold: float,
    files: Sequence[str],
) -> None:
    overview = summarize_overall(clusters)
    print_overall_summary(overview, files, threshold)
    print_cluster_details(clusters)


# --------------------------------------------------------------------------------------
# CLI
# --------------------------------------------------------------------------------------

def parse_args(argv: Optional[Sequence[str]] = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Agrupa mensagens (~70%% de similaridade) e imprime um relatorio completo no terminal.",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    parser.add_argument(
        "inputs",
        metavar="ARQUIVO",
        nargs="+",
        help="Arquivos ou padroes glob (ex: relatorio-*.xlsx).",
    )
    parser.add_argument(
        "--aba",
        dest="sheet_name",
        help="Nome da aba a ser utilizada. Se omitido, usa a primeira aba.",
    )
    parser.add_argument(
        "--saida",
        dest="legacy_output",
        help="Parametro legado (ignorado). O relatorio agora e exibido diretamente no terminal.",
    )
    parser.add_argument(
        "--threshold",
        type=float,
        default=0.70,
        help="Similaridade minima (0-1) para considerar dois corpos como pertencentes ao mesmo grupo.",
    )
    return parser.parse_args(argv)


def main(argv: Optional[Sequence[str]] = None) -> int:
    args = parse_args(argv)
    try:
        files = discover_input_files(args.inputs)
    except FileNotFoundError as exc:
        sys.stderr.write(f"{exc}\n")
        return 1

    if not (0 < args.threshold <= 1):
        sys.stderr.write("[ERRO] Threshold deve estar entre 0 e 1.\n")
        return 1

    messages = load_messages(files, sheet_name=args.sheet_name)
    if not messages:
        sys.stderr.write("[ERRO] Nenhuma mensagem valida encontrada nas planilhas informadas.\n")
        return 1

    clusters = cluster_messages(messages, threshold=args.threshold)
    if not clusters:
        sys.stderr.write("[ERRO] Falha ao gerar agrupamentos.\n")
        return 1
    if getattr(args, "legacy_output", None):
        sys.stderr.write("[AVISO] Parametro --saida ignorado: a saida agora e impressa no terminal.\n")

    print_console_report(clusters, threshold=args.threshold, files=files)
    print(f"[OK] Relatorio gerado no terminal.")
    print(f"   Arquivos de entrada: {', '.join(files)}")
    print(f"   Agrupamentos identificados: {len(clusters)}")
    return 0


if __name__ == "__main__":  # pragma: no cover
    raise SystemExit(main())
