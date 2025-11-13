#!/usr/bin/env python3
# python conerte-csv.py relatorio-1838922.xlsx --delimiter=';' --progress


from __future__ import annotations

import argparse
import csv
import datetime as dt
import sys
from pathlib import Path
from typing import Iterable, Optional

try:
    from openpyxl import load_workbook
    from openpyxl.worksheet.worksheet import Worksheet
except ImportError:  # pragma: no cover - executed only when dependency is missing
    raise SystemExit(
        "Missing dependency: openpyxl. Install it with 'pip install openpyxl'."
    )


DATE_FMT = "%d/%m/%Y"
TIME_FMT = "%H:%M:%S"
DATETIME_FMT = "%d/%m/%Y %H:%M:%S"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Converts a large XLSX file to CSV without loading the entire sheet into memory."
        )
    )
    parser.add_argument(
        "xlsx_path",
        type=Path,
        help="Path to the XLSX file to be converted.",
    )
    parser.add_argument(
        "csv_path",
        nargs="?",
        type=Path,
        help="Path to the CSV output file (defaults to <xlsx>.csv).",
    )
    parser.add_argument(
        "--sheet",
        help="Worksheet name or zero-based index to export (defaults to the active sheet).",
    )
    parser.add_argument(
        "--delimiter",
        default=";",
        help="Delimiter to use in the CSV file (default: ';').",
    )
    parser.add_argument(
        "--encoding",
        default="utf-8",
        help="Encoding for the CSV output (default: utf-8).",
    )
    parser.add_argument(
        "--progress",
        action="store_true",
        help="Print progress info every 10000 rows.",
    )
    return parser.parse_args()


def resolve_sheet(workbook, sheet_arg: Optional[str]) -> Worksheet:
    if sheet_arg is None:
        return workbook.active

    # Try by name first
    if sheet_arg in workbook.sheetnames:
        return workbook[sheet_arg]

    # Fall back to index
    try:
        index = int(sheet_arg)
    except ValueError as exc:
        available = ", ".join(workbook.sheetnames)
        raise SystemExit(
            f"Sheet '{sheet_arg}' not found. Available sheets: {available}"
        ) from exc

    try:
        return workbook.worksheets[index]
    except IndexError as exc:
        available = ", ".join(workbook.sheetnames)
        raise SystemExit(
            f"Sheet index {index} is out of range. Available sheets: {available}"
        ) from exc


def format_cell(cell) -> str:
    value = cell.value
    if value is None:
        return ""

    if cell.is_date:
        fmt = cell.number_format.lower()
        if isinstance(value, dt.datetime):
            # Excel time-only cells may be stored as 1899-12-30/31 plus time.
            if not any(token in fmt for token in ("y", "d")) and any(
                token in fmt for token in ("h", "m", "s")
            ):
                return value.strftime(TIME_FMT)
            if value.time() == dt.time():
                return value.strftime(DATE_FMT)
            if value.date() in {
                dt.date(1899, 12, 30),
                dt.date(1899, 12, 31),
            }:
                return value.strftime(TIME_FMT)
            return value.strftime(DATETIME_FMT)
        if isinstance(value, dt.date):
            return value.strftime(DATE_FMT)
        if isinstance(value, dt.time):
            return value.strftime(TIME_FMT)

    return str(value)


def iter_rows(ws: Worksheet) -> Iterable[list[str]]:
    for row in ws.iter_rows(values_only=False):
        yield [format_cell(cell) for cell in row]


def convert_to_csv(
    xlsx_path: Path,
    csv_path: Path,
    sheet_arg: Optional[str],
    delimiter: str,
    encoding: str,
    progress: bool,
) -> None:
    workbook = load_workbook(
        filename=xlsx_path,
        read_only=True,
        data_only=True,
    )
    worksheet = resolve_sheet(workbook, sheet_arg)

    csv_path.parent.mkdir(parents=True, exist_ok=True)

    with csv_path.open("w", newline="", encoding=encoding) as csv_file:
        writer = csv.writer(csv_file, delimiter=delimiter)
        for idx, row in enumerate(iter_rows(worksheet), start=1):
            writer.writerow(row)
            if progress and idx % 10000 == 0:
                print(f"{idx:,} rows written...", file=sys.stderr)

    workbook.close()


def main() -> None:
    args = parse_args()

    if not args.xlsx_path.exists():
        raise SystemExit(f"File not found: {args.xlsx_path}")

    csv_path = args.csv_path
    if csv_path is None:
        csv_path = args.xlsx_path.with_suffix(".csv")

    convert_to_csv(
        xlsx_path=args.xlsx_path,
        csv_path=csv_path,
        sheet_arg=args.sheet,
        delimiter=args.delimiter,
        encoding=args.encoding,
        progress=args.progress,
    )


if __name__ == "__main__":
    main()
