#!/usr/bin/env python3
import json
import re
import os
import csv
import glob
import textwrap
from collections import Counter
from pathlib import Path
from typing import Optional
import matplotlib.pyplot as plt
import matplotlib.patches as patches
from matplotlib.backends.backend_pdf import PdfPages
from datetime import datetime
from openpyxl import load_workbook

# Utilitário: formata nomes de arquivos para títulos por data
def _titulo_por_data(nome_ou_caminho: str) -> str:
    """Converte um nome de arquivo para um rótulo de data legível.

    - Remove a extensão
    - Tenta extrair datas nos formatos: ddmm, ddmmaaaa, dd-mm-aaaa, aaaa-mm-dd
    - Se não identificar, devolve o nome sem extensão/ruídos
    """
    base = os.path.basename(nome_ou_caminho)
    base = os.path.splitext(base)[0]

    # Normaliza separadores
    normalizado = base.replace('_', '-').replace('.', '-')

    # aaaa-mm-dd
    m = re.search(r"(20\d{2})[-_/]?([01]?\d)[-_/]?([0-3]?\d)", normalizado)
    if m:
        a, mth, d = m.groups()
        return f"{int(d):02d}/{int(mth):02d}/{a}"

    # dd-mm-aaaa
    m = re.search(r"([0-3]?\d)[-_/]?([01]?\d)[-_/]?(20\d{2})", normalizado)
    if m:
        d, mth, a = m.groups()
        return f"{int(d):02d}/{int(mth):02d}/{a}"

    # ddmmaaaa (ex.: 07092025)
    m = re.fullmatch(r"\D*(\d{2})(\d{2})(\d{4})\D*", re.sub(r"[^\d]", "", normalizado))
    if m:
        d, mth, a = m.groups()
        return f"{d}/{mth}/{a}"

    # ddmm (ex.: 0709) — sem ano
    m = re.fullmatch(r"\D*(\d{2})(\d{2})\D*", re.sub(r"[^\d]", "", normalizado))
    if m:
        d, mth = m.groups()
        return f"{d}/{mth}"

    # Fallback: devolve nome limpo
    return base

# ==============================
# CONFIGURAÇÕES - ALTERE AQUI 👇
# ==============================

# Lista de arquivos para analisar (JSON, JSONL ou CSV)
# Exemplos:
# ARQUIVOS = ["0709.csv"]                           # Um arquivo apenas
# ARQUIVOS = ["0709.csv", "0810.csv", "0911.csv"]  # Múltiplos arquivos específicos
# ARQUIVOS = ["*.csv"]                              # Todos os arquivos CSV no diretório
# ARQUIVOS = ["dados_*.csv"]                        # Arquivos que começam com "dados_"
ARQUIVOS = ["relatorio-hiper.xlsx"]

# Texto que deseja filtrar nas mensagens (case-insensitive).
# Se não quiser filtro, deixe vazio: FILTRO = ""
# Para usar regex, defina USAR_REGEX = True
# Filtro que captura o texto alvo permitindo variação no final da URL
FILTRO = r""
USAR_REGEX = True

# Informações do relatório
ID_FLOW = ""
TEXTO_SMS = (
    ""
)
DATA_RELATORIO = datetime.now().strftime("%d/%m/%Y")
ARQUIVO_PDF_SAIDA = f"relatorio_apostatudo_{datetime.now().strftime('%d%m%Y')}.pdf"
SUBTITLE = "" + ID_FLOW

MESSAGE_FIELDS = (
    "body",
    "message",
    "mensagem",
    "texto",
    "copy",
    "msg",
    "content",
    "message_body",
)

# ==============================
# FIM DAS CONFIGURAÇÕES
# ==============================

def _normalizar_texto_sms(valor: Optional[str]) -> Optional[str]:
    if valor is None:
        return None
    texto = str(valor).strip()
    return texto or None


def _extrair_copy_de_registro(registro: dict) -> Optional[str]:
    for campo in MESSAGE_FIELDS:
        if campo in registro:
            texto = _normalizar_texto_sms(registro.get(campo))
            if texto:
                return texto
    return None


def _resolver_texto_sms(texto_preferencial: Optional[str], filtro: Optional[str], resultados_por_arquivo):
    # Preferência explícita (campo vindo da GUI/CLI)
    texto = _normalizar_texto_sms(texto_preferencial)
    if texto:
        return texto

    # Em seguida, reutiliza o filtro informado
    texto = _normalizar_texto_sms(filtro)
    if texto:
        return texto

    # Por fim, pega o primeiro texto de mensagem encontrado nos dados filtrados
    for resultado in resultados_por_arquivo:
        registros = resultado.get("registros_filtrados") or []
        for registro in registros:
            texto_registro = _extrair_copy_de_registro(registro)
            if texto_registro:
                return texto_registro
    return None


def extrair_chave_telefone(registro):
    """Extrai uma chave de telefone normalizada para deduplicação."""
    campos_telefone = (
        "to", "To", "TO",
        "destination", "Destination",
        "destino", "telefone", "phone"
    )

    for campo in campos_telefone:
        valor = registro.get(campo)
        if not valor:
            continue

        texto = str(valor).strip()
        if not texto:
            continue

        digitos = re.sub(r"\D", "", texto)
        if digitos:
            return digitos
        return texto

    return None

def formatar_mensagem_para_pdf(texto: str, largura: int = 90) -> str:
    """Quebra o texto em múltiplas linhas adequadas ao PDF."""
    linhas = []
    for bloco in texto.splitlines():
        bloco = bloco.strip()
        if not bloco:
            linhas.append("")
            continue
        linhas.append(textwrap.fill(bloco, width=largura))
    return "\n".join(linhas)

def desenhar_cartoes_metricas(ax, metricas, titulo: str = None):
    """Renderiza cartões de métricas com um estilo consistente."""
    ax.set_xlim(0, 1)
    ax.set_ylim(0, 1)
    ax.axis('off')

    if not metricas:
        return

    if titulo:
        ax.text(0.02, 0.95, titulo, fontsize=12, fontweight='bold',
                color='#212529', va='top')

    espacamento = 0.02
    largura = (0.96 - espacamento * (len(metricas) - 1)) / len(metricas)
    y_base = 0.18
    altura = 0.64

    for idx, (rotulo, valor, cor) in enumerate(metricas):
        x = 0.02 + idx * (largura + espacamento)

        card = patches.FancyBboxPatch(
            (x, y_base),
            largura,
            altura,
            boxstyle="round,pad=0.02",
            facecolor="#f8f9fa",
            edgecolor="#dee2e6",
            linewidth=1.0
        )
        card.set_zorder(0.5)
        ax.add_patch(card)

        destaque = patches.Rectangle(
            (x, y_base + altura - 0.03),
            largura,
            0.018,
            color=cor,
            alpha=0.35,
            clip_on=False
        )
        destaque.set_zorder(0.6)
        ax.add_patch(destaque)

        ax.text(
            x + largura / 2,
            y_base + altura * 0.58,
            valor,
            ha='center',
            va='center',
            fontsize=14,
            fontweight='bold',
            color=cor
        )
        ax.text(
            x + largura / 2,
            y_base + altura * 0.28,
            rotulo,
            ha='center',
            va='center',
            fontsize=10.5,
            color='#495057'
        )

def descobrir_arquivos(lista_arquivos):
    """Descobre arquivos baseado na lista ou padrão glob"""
    arquivos_encontrados = []
    
    for item in lista_arquivos:
        if '*' in item or '?' in item:
            # É um padrão glob
            arquivos_glob = glob.glob(item)
            arquivos_encontrados.extend(sorted(arquivos_glob))
        else:
            # É um arquivo específico
            if os.path.exists(item):
                arquivos_encontrados.append(item)
            else:
                print(f"⚠️ Arquivo não encontrado: {item}")
    
    return arquivos_encontrados

def ler_arquivo(caminho_arquivo):
    """Lê arquivo JSON, JSONL ou CSV e retorna lista de dicionários"""
    
    # Detectar extensão do arquivo
    _, extensao = os.path.splitext(caminho_arquivo.lower())
    
    print(f"📄 Tipo de arquivo detectado: {extensao}")
    
    if extensao == '.csv':
        return ler_csv(caminho_arquivo)
    elif extensao == '.json':
        return ler_json(caminho_arquivo)
    elif extensao in ('.xlsx', '.xlsm', '.xltx', '.xltm'):
        return ler_xlsx(caminho_arquivo)
    else:
        # Tentar como JSON primeiro, depois XLSX e por fim CSV
        try:
            return ler_json(caminho_arquivo)
        except Exception:
            try:
                return ler_xlsx(caminho_arquivo)
            except Exception:
                print("📝 Tentando ler como CSV...")
                return ler_csv(caminho_arquivo)

def ler_xlsx(caminho_arquivo):
    """Lê arquivo XLSX e retorna lista de dicionários"""
    dados = []
    try:
        wb = load_workbook(caminho_arquivo, read_only=True, data_only=True)
        try:
            sheet = wb.active
            header = None
            for row_num, row in enumerate(sheet.iter_rows(values_only=True), 1):
                valores = list(row)

                # Identifica a linha de cabeçalho (primeira linha não vazia)
                if header is None:
                    if not any(
                        valor is not None and str(valor).strip()
                        for valor in valores
                    ):
                        continue

                    header = []
                    for idx, valor in enumerate(valores):
                        nome = str(valor).strip() if valor is not None else ""
                        if not nome:
                            nome = f"col_{idx+1}"
                        header.append(nome)
                    continue

                # Ignora linhas completamente vazias
                if not any(
                    valor is not None and (not isinstance(valor, str) or valor.strip())
                    for valor in valores
                ):
                    continue

                registro = {}
                for idx, coluna in enumerate(header):
                    valor = valores[idx] if idx < len(valores) else None

                    if hasattr(valor, "isoformat"):
                        try:
                            valor = valor.isoformat()
                        except TypeError:
                            valor = str(valor)
                    elif isinstance(valor, str):
                        valor = valor.strip()

                    registro[coluna] = valor

                dados.append(registro)

                if len(dados) % 1000 == 0:
                    print(f"📖 Lendo linha {row_num}...")
        finally:
            wb.close()

        if header is None:
            print("⚠️ Nenhum cabeçalho encontrado no XLSX.")
        print(f"✅ XLSX lido com {len(dados)} registros.")
        return dados
    except Exception as e:
        print(f"❌ Erro ao ler XLSX: {e}")
        return []

def ler_csv(caminho_arquivo):
    """Lê arquivo CSV e retorna lista de dicionários"""
    dados = []
    
    try:
        with open(caminho_arquivo, 'r', encoding='utf-8') as file:
            # Detectar delimitador automaticamente
            sample = file.read(1024)
            file.seek(0)
            
            # Tentar diferentes delimitadores
            delimitadores = [',', ';', '\t', '|']
            melhor_delimitador = ','
            
            for delim in delimitadores:
                if sample.count(delim) > sample.count(melhor_delimitador):
                    melhor_delimitador = delim
            
            print(f"📊 Delimitador CSV detectado: '{melhor_delimitador}'")
            
            reader = csv.DictReader(file, delimiter=melhor_delimitador)
            
            for linha_num, linha in enumerate(reader, 2):  # Começa em 2 pois linha 1 é header
                if linha:  # Ignora linhas vazias
                    dados.append(linha)
                
                # Mostrar progresso a cada 1000 linhas
                if linha_num % 1000 == 0:
                    print(f"📖 Lendo linha {linha_num}...")
        
        print(f"✅ CSV lido com {len(dados)} registros.")
        return dados
        
    except Exception as e:
        print(f"❌ Erro ao ler CSV: {e}")
        return []

def ler_json(caminho_arquivo):
    """Lê arquivo JSON ou JSONL e retorna lista de dicionários"""
    try:
        with open(caminho_arquivo, 'r', encoding='utf-8') as file:
            dados = json.load(file)
            
        # Converte para lista se for um único objeto
        if isinstance(dados, dict):
            dados = [dados]
            
        print(f"✅ JSON lido com {len(dados)} registros.")
        return dados
        
    except json.JSONDecodeError as e:
        print(f"❌ Erro ao ler o arquivo JSON: {e.msg}")
        print(f"📄 Tentando ler como JSONL (JSON Lines)...")
        
        # Tenta ler como JSONL (uma linha JSON por linha)
        try:
            dados = []
            with open(caminho_arquivo, 'r', encoding='utf-8') as file:
                for line_num, line in enumerate(file, 1):
                    line = line.strip()
                    if line:  # Ignora linhas vazias
                        try:
                            dados.append(json.loads(line))
                        except json.JSONDecodeError as line_error:
                            print(f"❌ Erro na linha {line_num}: {line_error.msg}")
                            continue
                    
                    # Mostrar progresso a cada 1000 linhas
                    if line_num % 1000 == 0:
                        print(f"📖 Lendo linha {line_num}...")
            
            if dados:
                print(f"✅ Arquivo lido como JSONL com {len(dados)} registros.")
                return dados
            else:
                print(f"❌ Nenhum registro válido encontrado.")
                return []
                
        except Exception as e:
            print(f"❌ Erro ao tentar ler como JSONL: {e}")
        return []

def analisar_dados(dados, nome_arquivo=""):
    """Analisa os dados e retorna estatísticas"""
    if nome_arquivo:
        print(f"📊 Analisando arquivo: {nome_arquivo}")
    print(f"📊 Total de registros no arquivo: {len(dados)}")
    
    # Contadores
    status_counter = Counter()
    registros_filtrados = []
    total_processados = 0
    
    for registro in dados:
        total_processados += 1
        
        # Pega o status e body
        status = registro.get('status', 'sem_status')
        body_text = registro.get('body', '')
        
        # Aplica o filtro se definido
        if FILTRO and body_text:
            if USAR_REGEX:
                if not re.search(FILTRO, body_text, re.IGNORECASE):
                    continue
            else:
                if FILTRO.lower() not in body_text.lower():
                    continue
        elif FILTRO and not body_text:
            # Se tem filtro mas não tem body, pula
            continue
        
        # Registro passou no filtro
        registros_filtrados.append(registro)
        status_counter[status] += 1
        
        # Print de progresso a cada 1000 registros
        if total_processados % 1000 == 0:
            print(f"📊 Processados: {total_processados} | Filtrados: {len(registros_filtrados)}")
    
    print("─" * 40)
    print(f"✅ Total filtrado: {len(registros_filtrados)} registros")
    
    return {
        'status_counter': dict(status_counter),
        'total_filtrado': len(registros_filtrados),
        'total_processado': total_processados,
        'registros_filtrados': registros_filtrados,
        'nome_arquivo': nome_arquivo
    }

def analisar_multiplos_arquivos(lista_arquivos):
    """Analisa múltiplos arquivos e retorna dados consolidados"""
    resultados_por_arquivo = []
    dados_consolidados = {
        'status_counter': Counter(),
        'total_filtrado': 0,
        'total_processado': 0,
        'arquivos_processados': 0
    }
    
    print(f"🎯 Filtro regex ativo: {FILTRO}")
    print("=" * 80)
    
    for arquivo in lista_arquivos:
        print(f"\n� Processando arquivo: {arquivo}")
        print("─" * 40)
        
        # Ler arquivo
        dados = ler_arquivo(arquivo)
        
        if not dados:
            print(f"❌ Nenhum dado carregado de {arquivo}")
            continue
        
        # Analisar dados do arquivo
        resultado = analisar_dados(dados, arquivo)
        resultados_por_arquivo.append(resultado)
        
        # Consolidar dados
        for status, quantidade in resultado['status_counter'].items():
            dados_consolidados['status_counter'][status] += quantidade
        
        dados_consolidados['total_filtrado'] += resultado['total_filtrado']
        dados_consolidados['total_processado'] += resultado['total_processado']
        dados_consolidados['arquivos_processados'] += 1
        
        print(f"📈 Resumo do arquivo {arquivo}:")
        for status, quantidade in resultado['status_counter'].items():
            porcentagem = (quantidade / resultado['total_filtrado'] * 100) if resultado['total_filtrado'] > 0 else 0
            print(f"   📊 {status:15} → {quantidade:4} registros ({porcentagem:5.1f}%)")
    
    print("\n" + "=" * 80)
    print("🎉 ANÁLISE CONSOLIDADA:")
    print("=" * 80)
    print(f"📁 Arquivos processados: {dados_consolidados['arquivos_processados']}")
    print(f"� Total de registros processados: {dados_consolidados['total_processado']:,}")
    print(f"✅ Total de mensagens filtradas: {dados_consolidados['total_filtrado']:,}")
    print()
    
    if dados_consolidados['total_filtrado'] > 0:
        print("📈 RESUMO CONSOLIDADO POR STATUS:")
        print("─" * 50)
        
        for status, quantidade in dados_consolidados['status_counter'].most_common():
            porcentagem = (quantidade / dados_consolidados['total_filtrado']) * 100
            print(f"📊 {status:15} → {quantidade:6,} registros ({porcentagem:5.1f}%)")
        
        print("─" * 50)
        print(f"📋 Total consolidado: {dados_consolidados['total_filtrado']:,} mensagens filtradas")
        print()
    
    return dict(dados_consolidados), resultados_por_arquivo

def criar_pagina_graficos(dados_status):
    """Cria uma página com dois gráficos lado a lado"""
    # Configurar figura A4 retrato (8.27 x 11.69 polegadas)
    fig = plt.figure(figsize=(8.27, 11.69), dpi=100)
    
    # Cores personalizadas
    cores = {
        'delivered': '#28a745',     # Verde
        'undelivered': '#ffc107',   # Amarelo
        'failed': '#dc3545',        # Vermelho
        'sent': '#17a2b8'           # Azul
    }
    
    total = sum(dados_status.values())
    
    # GRÁFICO DE PIZZA (lado esquerdo)
    ax1 = fig.add_subplot(121)
    
    labels = []
    sizes = []
    colors = []
    
    for status, quantidade in dados_status.items():
        porcentagem = (quantidade / total) * 100
        labels.append(f'{status.title()}\n{quantidade:,}\n({porcentagem:.1f}%)')
        sizes.append(quantidade)
        colors.append(cores.get(status, '#6c757d'))
    
    # Criar gráfico de pizza
    wedges, texts, autotexts = ax1.pie(sizes, labels=labels, colors=colors, 
                                       autopct='', startangle=90, 
                                       textprops={'fontsize': 10},
                                       labeldistance=1.15,
                                       radius=0.9)
    
    ax1.set_title('Distribuição por Status', fontsize=14, fontweight='bold', pad=20)
    
    # GRÁFICO DE BARRAS (lado direito)
    ax2 = fig.add_subplot(122)
    
    status_list = list(dados_status.keys())
    valores = list(dados_status.values())
    cores_barras = [cores.get(status, '#6c757d') for status in status_list]
    
    # Criar barras
    bars = ax2.bar(status_list, valores, color=cores_barras, edgecolor='black', linewidth=1)
    
    # Adicionar valores nas barras
    for bar, valor in zip(bars, valores):
        height = bar.get_height()
        ax2.text(bar.get_x() + bar.get_width()/2., height + max(valores)*0.01,
                f'{valor:,}', ha='center', va='bottom', fontsize=10, fontweight='bold')
    
    ax2.set_title('Quantidade por Status', fontsize=14, fontweight='bold', pad=20)
    ax2.set_ylabel('Quantidade de Mensagens', fontsize=12)
    ax2.set_xlabel('Status', fontsize=12)
    
    # Rotacionar labels do eixo x se necessário
    ax2.tick_params(axis='x', rotation=45)
    
    # Ajustar layout
    plt.tight_layout()
    
    # Adicionar título geral
    # fig.suptitle('Análise de Status - Fortune Dragon HiperBet', 
    #              fontsize=16, fontweight='bold', y=0.95)
    
    return fig
    ax = fig.add_subplot(111)
    
    # Cores personalizadas
    cores = {
        'delivered': '#28a745',
        'undelivered': '#ffc107', 
        'failed': '#dc3545',
        'sent': '#17a2b8'
    }
    
    status_list = list(dados_status.keys())
    quantidades = list(dados_status.values())
    cores_list = [cores.get(status, '#6c757d') for status in status_list]
    
    # Criar barras com largura adequada para retrato
    bars = ax.bar(status_list, quantidades, color=cores_list, alpha=0.8, width=0.6)
    
    # Adicionar valores nas barras
    for bar, quantidade in zip(bars, quantidades):
        height = bar.get_height()
        ax.text(bar.get_x() + bar.get_width()/2., height + height*0.01,
                f'{quantidade:,}', ha='center', va='bottom', fontweight='bold', fontsize=12)
    
    # Título
    ax.set_title('Quantidade de Mensagens por Status\nFortune Dragon - HiperBet', 
                 fontsize=18, fontweight='bold', pad=30)
    ax.set_ylabel('Quantidade de Mensagens', fontsize=12, fontweight='bold')
    ax.set_xlabel('Status', fontsize=12, fontweight='bold')
    
    # Formatar eixo Y
    ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: f'{x:,.0f}'))
    
    # Ajustar fontes e aparência
    ax.tick_params(axis='both', which='major', labelsize=12)
    
    # Grid sutil
    ax.grid(True, alpha=0.3, axis='y')
    ax.set_axisbelow(True)
    
    # Ajustar limites para não cortar
    ax.margins(x=0.1, y=0.1)
    
    # Margens específicas para A4 retrato
    fig.subplots_adjust(left=0.15, right=0.9, top=0.85, bottom=0.15)
    
    return fig

def criar_pagina_resumo(dados_status, total_filtrado, total_processado):
    """Cria página com resumo executivo com layout compacto"""
    # Tamanho A4 retrato (8.27 x 11.69 polegadas)
    fig = plt.figure(figsize=(8.27, 11.69), dpi=100)
    ax = fig.add_subplot(111)
    ax.axis('off')
    
    total = sum(dados_status.values())
    delivered = dados_status.get('delivered', 0)
    taxa_sucesso = (delivered / total) * 100 if total > 0 else 0
    
    # Título principal
    ax.text(0.5, 0.95, 'RELATÓRIO DE CAMPANHAS SMS', 
            fontsize=16, fontweight='bold', ha='center', va='top')
    
    # Subtítulo
    # ax.text(0.5, 0.91, SUBTITLE, 
    #         fontsize=14, ha='center', va='top', color='#2c3e50')
    
    # Linha separadora
    ax.plot([0.1, 0.9], [0.87, 0.87], 'k-', linewidth=2)
    
    # Informações gerais em colunas
    y_pos = 0.83
    ax.text(0.1, y_pos, 'INFORMAÇÕES DA CAMPANHA:', fontsize=14, fontweight='bold', color='#2c3e50')
    
    y_pos -= 0.04
    ax.text(0.1, y_pos, f'• ID do Flow:', fontsize=11, fontweight='bold')
    ax.text(0.35, y_pos, f'{ID_FLOW}', fontsize=11)
    
    # Removida a exibição de data do relatório
    
    # Total Enviado (sem Total Processado)
    y_pos -= 0.06
    ax.text(0.1, y_pos, f'• Total Enviado:', fontsize=11, fontweight='bold')
    ax.text(0.35, y_pos, f'{total_filtrado:,} mensagens', fontsize=11)
    
    y_pos -= 0.03
    ax.text(0.1, y_pos, f'• Taxa de Sucesso:', fontsize=11, fontweight='bold')
    ax.text(0.35, y_pos, f'{taxa_sucesso:.1f}%', fontsize=11, fontweight='bold', color='#28a745')
    
    # Linha separadora
    y_pos -= 0.03
    ax.plot([0.1, 0.9], [y_pos, y_pos], 'k-', linewidth=1)
    
    # Resumo dos status em formato mais compacto
    y_pos -= 0.04
    ax.text(0.1, y_pos, 'RESUMO DOS STATUS:', fontsize=14, fontweight='bold', color='#2c3e50')
    
    cores = {
        'delivered': '#28a745',
        'undelivered': '#ffc107',
        'failed': '#dc3545',
        'sent': '#17a2b8'
    }
    
    # Criar grid de status 2x2
    status_items = list(dados_status.items())
    col1_x, col2_x = 0.1, 0.5
    
    for i, (status, quantidade) in enumerate(status_items):
        if i < 2:  # Primeira coluna
            x_pos = col1_x
            y_offset = i * 0.05
        else:  # Segunda coluna
            x_pos = col2_x
            y_offset = (i - 2) * 0.05
        
        y_status = y_pos - 0.04 - y_offset
        porcentagem = (quantidade / total) * 100
        
        # Criar retângulo colorido
        rect = patches.Rectangle((x_pos, y_status-0.01), 0.02, 0.02, 
                               facecolor=cores.get(status, '#6c757d'))
        ax.add_patch(rect)
        
        ax.text(x_pos + 0.03, y_status, f'{status.title()}: {quantidade:,} ({porcentagem:.1f}%)', 
                fontsize=11, va='center', fontweight='bold')
    
    # Texto da mensagem
    y_pos -= 0.18
    ax.text(0.1, y_pos, 'TEXTO DA MENSAGEM:', fontsize=14, fontweight='bold', color='#2c3e50')
    
    y_pos -= 0.04
    # Quebrar o texto em linhas para melhor visualização
    texto_quebrado = formatar_mensagem_para_pdf(TEXTO_SMS, largura=80)
    ax.text(0.1, y_pos, texto_quebrado, fontsize=10, va='top', 
            bbox=dict(boxstyle="round,pad=0.5", facecolor="#f8f9fa", edgecolor="#dee2e6"))
    
    return fig

def criar_pagina_consolidada(dados_consolidados, resultados_por_arquivo):
    """Cria página consolidada com dados de todos os arquivos"""
    
    # Criar figura A4 retrato
    fig = plt.figure(figsize=(8.27, 11.69))  # A4 retrato em polegadas
    
    status_counter = dados_consolidados['status_counter']
    total_filtrado = dados_consolidados['total_filtrado']
    total_processado = dados_consolidados['total_processado']
    
    # Título principal e mensagem da campanha
    datas = [_titulo_por_data(r['nome_arquivo']) for r in resultados_por_arquivo]
    periodo = ", ".join(datas)
    fig.suptitle('RELATÓRIO CONSOLIDADO - CAMPANHA SMS', fontsize=16, fontweight='bold', y=0.985)
    
    mensagem_formatada = formatar_mensagem_para_pdf(TEXTO_SMS, largura=92)
    fig.text(
        0.5,
        0.93,
        f"Mensagem da campanha:\n{mensagem_formatada}",
        ha='center',
        va='top',
        fontsize=11,
        color='#343a40',
        linespacing=1.35,
        bbox=dict(boxstyle="round,pad=0.6", facecolor="#f8f9fa", edgecolor="#dee2e6")
    )

    # Indicadores principais
    total_entregues = status_counter.get('delivered', 0)
    falhas = status_counter.get('failed', 0)
    taxa_sucesso = (total_entregues / total_filtrado * 100) if total_filtrado > 0 else 0
    metricas = [
        ("Mensagens enviadas", f"{total_filtrado:,}", '#0d6efd'),
        ("Entregues", f"{total_entregues:,}", '#198754'),
        ("Taxa de sucesso", f"{taxa_sucesso:.1f}%", '#fd7e14'),
        ("Falhas", f"{falhas:,}", '#dc3545'),
    ]
    ax_resumo = fig.add_axes([0.08, 0.60, 0.84, 0.18])
    desenhar_cartoes_metricas(ax_resumo, metricas, titulo="Indicadores principais")
    
    # Paleta de cores consistente
    cores = {
        'delivered': '#28a745',
        'undelivered': '#ffc107',
        'failed': '#dc3545',
        'sent': '#17a2b8',
        'queued': '#6c757d'
    }
    
    # ====== GRÁFICO DE PIZZA (ESQUERDA) ======
    ax_pizza = fig.add_axes([0.08, 0.33, 0.38, 0.23])
    valores = list(status_counter.values())
    labels = list(status_counter.keys())
    cores_grafico = [cores.get(label, '#95a5a6') for label in labels]
    ax_pizza.set_facecolor("#ffffff")
    
    if valores:
        wedges, texts, autotexts = ax_pizza.pie(
            valores,
            labels=labels,
            autopct='%1.1f%%',
            colors=cores_grafico,
            startangle=90
        )
        ax_pizza.axis('equal')
        for text in texts:
            text.set_fontsize(8)
        for autotext in autotexts:
            autotext.set_fontsize(7)
            autotext.set_fontweight('bold')
            autotext.set_color('white')
    ax_pizza.set_title('Distribuição por status', fontsize=12, fontweight='bold', pad=10)
    
    # ====== GRÁFICO DE BARRAS (DIREITA) ======
    ax_barras = fig.add_axes([0.54, 0.33, 0.38, 0.23])
    ax_barras.set_facecolor("#ffffff")
    valores_b = list(status_counter.values())
    labels_b = list(status_counter.keys())
    cores_barras = [cores.get(label, '#95a5a6') for label in labels_b]
    
    if valores_b:
        y_pos_b = range(len(labels_b))
        bars = ax_barras.barh(y_pos_b, valores_b, color=cores_barras)
        ax_barras.set_yticks(y_pos_b)
        ax_barras.set_yticklabels(labels_b)
        ax_barras.set_xlabel('Quantidade', fontsize=9)
        ax_barras.set_title('Quantidades por status', fontsize=12, fontweight='bold')
        for bar, val in zip(bars, valores_b):
            ax_barras.text(
                bar.get_width() + max(valores_b) * 0.01,
                bar.get_y() + bar.get_height() / 2,
                f'{val:,}',
                va='center',
                fontsize=8,
                fontweight='bold'
            )
        ax_barras.tick_params(axis='both', which='major', labelsize=8)
        ax_barras.grid(True, axis='x', alpha=0.2, linestyle='--')
    for spine in ax_barras.spines.values():
        spine.set_visible(False)
    
    # ====== DETALHAMENTO DOS STATUS CONSOLIDADO ======
    ax_detalhes = fig.add_axes([0.08, 0.07, 0.84, 0.22])
    ax_detalhes.set_xlim(0, 1)
    ax_detalhes.set_ylim(0, 1)
    ax_detalhes.axis('off')
    
    bloco = patches.FancyBboxPatch(
        (0.0, 0.0), 1, 1,
        boxstyle="round,pad=0.03",
        facecolor="#fdfdfd",
        edgecolor="#e9ecef",
        linewidth=1.0
    )
    bloco.set_zorder(0.4)
    ax_detalhes.add_patch(bloco)
    
    ax_detalhes.text(
        0.5, 0.88,
        'Detalhamento consolidado dos status',
        fontsize=13,
        fontweight='bold',
        ha='center',
        color='#2c3e50'
    )
    
    status_descricoes = {
        'delivered': 'Mensagens entregues com sucesso',
        'undelivered': 'Mensagens não entregues',
        'falhas': 'Mensagens que falharam no envio',
        'sent': 'Mensagens enviadas (aguardando confirmação)',
        'queued': 'Mensagens na fila de envio'
    }
    
    y_start = 0.74
    passo = 0.16
    for i, (status, quantidade) in enumerate(status_counter.items()):
        y_pos = y_start - (i * passo)
        cor = cores.get(status, '#95a5a6')
        porcentagem = (quantidade / total_filtrado * 100) if total_filtrado > 0 else 0
        descricao = status_descricoes.get(status, 'Status desconhecido')
        
        marcador = patches.Rectangle(
            (0.05, y_pos - 0.015),
            0.02,
            0.03,
            facecolor=cor,
            edgecolor='none'
        )
        marcador.set_zorder(0.6)
        ax_detalhes.add_patch(marcador)
        ax_detalhes.text(
            0.08,
            y_pos,
            f'{status.upper()}: {quantidade:,} mensagens ({porcentagem:.1f}%) - {descricao}',
            fontsize=10.5,
            fontweight='bold',
            color='#343a40',
            va='center'
        )
    
    return fig

def criar_pagina_individual(resultado_arquivo):
    """Cria página individual para um arquivo específico"""
    
    # Título baseado em data (sem extensão)
    nome_arquivo = _titulo_por_data(resultado_arquivo['nome_arquivo'])
    dados_status = resultado_arquivo['status_counter']
    total_filtrado = resultado_arquivo['total_filtrado']
    total_processado = resultado_arquivo['total_processado']
    
    # Criar figura A4 retrato
    fig = plt.figure(figsize=(8.27, 11.69))
    
    # Título principal com ênfase em data
    fig.suptitle(f'RELATÓRIO POR DATA - {nome_arquivo}', fontsize=16, fontweight='bold', y=0.985)
    
    mensagem_formatada = formatar_mensagem_para_pdf(TEXTO_SMS, largura=92)
    fig.text(
        0.5,
        0.93,
        f"Mensagem da campanha:\n{mensagem_formatada}",
        ha='center',
        va='top',
        fontsize=11,
        color='#343a40',
        linespacing=1.35,
        bbox=dict(boxstyle="round,pad=0.6", facecolor="#f8f9fa", edgecolor="#dee2e6")
    )
    origem = os.path.basename(resultado_arquivo['nome_arquivo'])
    fig.text(0.5, 0.865, f'Arquivo analisado: {origem}', ha='center', fontsize=10, color='#6c757d')
    
    # Calcular porcentagens e taxa de sucesso
    total_entregues = dados_status.get('delivered', 0)
    falhas = dados_status.get('failed', 0)
    taxa_sucesso = (total_entregues / total_filtrado * 100) if total_filtrado > 0 else 0
    
    # Indicadores principais
    metricas = [
        ("Mensagens enviadas", f"{total_filtrado:,}", '#0d6efd'),
        ("Entregues", f"{total_entregues:,}", '#198754'),
        ("Taxa de sucesso", f"{taxa_sucesso:.1f}%", '#fd7e14'),
        ("Failed", f"{falhas:,}", '#dc3545'),
    ]
    ax_resumo = fig.add_axes([0.08, 0.60, 0.84, 0.18])
    desenhar_cartoes_metricas(ax_resumo, metricas, titulo=f"Indicadores do dia {nome_arquivo}")
    
    cores = {
        'delivered': '#28a745',
        'undelivered': '#ffc107',
        'failed': '#dc3545',
        'sent': '#17a2b8',
        'queued': '#6c757d'
    }
    
    labels = list(dados_status.keys())
    valores = list(dados_status.values())
    cores_grafico = [cores.get(label, '#95a5a6') for label in labels]
    
    # ====== GRÁFICO DE PIZZA (ESQUERDA) ======
    ax_pizza = fig.add_axes([0.08, 0.33, 0.38, 0.23])
    ax_pizza.set_facecolor("#ffffff")
    if valores:
        wedges, texts, autotexts = ax_pizza.pie(
            valores,
            labels=labels,
            autopct='%1.1f%%',
            colors=cores_grafico,
            startangle=90
        )
        ax_pizza.axis('equal')
        for text in texts:
            text.set_fontsize(8)
        for autotext in autotexts:
            autotext.set_fontsize(7)
            autotext.set_fontweight('bold')
            autotext.set_color('white')
    ax_pizza.set_title('Distribuição por status', fontsize=12, fontweight='bold', pad=10)
    
    # ====== GRÁFICO DE BARRAS (DIREITA) ======
    ax_barras = fig.add_axes([0.54, 0.33, 0.38, 0.23])
    ax_barras.set_facecolor("#ffffff")
    if valores:
        y_pos_bar = range(len(labels))
        bars = ax_barras.barh(y_pos_bar, valores, color=cores_grafico)
        ax_barras.set_yticks(y_pos_bar)
        ax_barras.set_yticklabels(labels)
        ax_barras.set_xlabel('Quantidade', fontsize=9)
        ax_barras.set_title('Quantidades por status', fontsize=12, fontweight='bold')
        for bar, valor in zip(bars, valores):
            ax_barras.text(
                bar.get_width() + max(valores) * 0.01,
                bar.get_y() + bar.get_height() / 2,
                f'{valor:,}',
                va='center',
                fontsize=8,
                fontweight='bold'
            )
        ax_barras.tick_params(axis='both', which='major', labelsize=8)
        ax_barras.grid(True, axis='x', alpha=0.2, linestyle='--')
    for spine in ax_barras.spines.values():
        spine.set_visible(False)
    
    # ====== DETALHAMENTO DOS STATUS (PARTE INFERIOR) ======
    ax_detalhes = fig.add_axes([0.08, 0.07, 0.84, 0.22])
    ax_detalhes.set_xlim(0, 1)
    ax_detalhes.set_ylim(0, 1)
    ax_detalhes.axis('off')
    
    bloco = patches.FancyBboxPatch(
        (0.0, 0.0), 1, 1,
        boxstyle="round,pad=0.03",
        facecolor="#fdfdfd",
        edgecolor="#e9ecef",
        linewidth=1.0
    )
    bloco.set_zorder(0.4)
    ax_detalhes.add_patch(bloco)
    
    ax_detalhes.text(
        0.5, 0.88,
        f'Detalhamento do dia {nome_arquivo}',
        fontsize=13,
        fontweight='bold',
        ha='center',
        color='#2c3e50'
    )
    
    status_descricoes = {
        'delivered': 'Mensagens entregues com sucesso',
        'undelivered': 'Mensagens não entregues',
        'failed': 'Mensagens que falharam no envio',
        'sent': 'Mensagens enviadas (aguardando confirmação)',
        'queued': 'Mensagens na fila de envio'
    }
    
    y_start = 0.74
    passo = 0.16
    for i, (status, quantidade) in enumerate(dados_status.items()):
        y_pos = y_start - (i * passo)
        cor = cores.get(status, '#95a5a6')
        porcentagem = (quantidade / total_filtrado * 100) if total_filtrado > 0 else 0
        descricao = status_descricoes.get(status, 'Erro (exemplo: telefone inválido)')
        
        marcador = patches.Rectangle(
            (0.05, y_pos - 0.015),
            0.02,
            0.03,
            facecolor=cor,
            edgecolor='none'
        )
        marcador.set_zorder(0.6)
        ax_detalhes.add_patch(marcador)
        ax_detalhes.text(
            0.08,
            y_pos,
            f'{status.upper()}: {quantidade:,} mensagens ({porcentagem:.1f}%) - {descricao}',
            fontsize=10.5,
            fontweight='bold',
            color='#343a40',
            va='center'
        )
    
    return fig

def criar_pagina_completa(dados_status, total_filtrado, total_processado):
    
    # Criar figura A4 retrato
    fig = plt.figure(figsize=(8.27, 11.69))  # A4 retrato em polegadas
    
    # Título principal
    fig.suptitle('RELATÓRIO DE CAMPANHA SMS\nFortune Dragon - HiperBet', 
                 fontsize=18, fontweight='bold', y=0.97)
    
    # Calcular porcentagens e taxa de sucesso
    total_entregues = dados_status.get('delivered', 0)
    taxa_sucesso = (total_entregues / total_filtrado * 100) if total_filtrado > 0 else 0
    
    # ====== SEÇÃO DE RESUMO (TOPO) ======
    ax_resumo = fig.add_axes([0.1, 0.7, 0.8, 0.2])  # [left, bottom, width, height]
    ax_resumo.set_xlim(0, 1)
    ax_resumo.set_ylim(0, 1)
    ax_resumo.axis('off')
    
    # Título da seção
    ax_resumo.text(0.5, 0.9, 'RESUMO EXECUTIVO', fontsize=16, fontweight='bold', 
                   ha='center', color='#2c3e50')
    
    # Estatísticas principais em duas colunas
    y_pos = 0.7
    
    # Coluna 1 (removido Total Processado; Total Filtrado -> Total Enviado)
    ax_resumo.text(0.05, y_pos-0.1, f'Total Enviado: {total_filtrado:,}', fontsize=12, fontweight='bold')
    
    # Coluna 2
    ax_resumo.text(0.55, y_pos, f'Entregues: {total_entregues:,}', fontsize=12, fontweight='bold', color='#28a745')
    ax_resumo.text(0.55, y_pos-0.1, f'Taxa de Sucesso: {taxa_sucesso:.1f}%', fontsize=12, 
                   color='#28a745', fontweight='bold')
    
    # Linha separadora
    ax_resumo.plot([0.05, 0.95], [y_pos-0.25, y_pos-0.25], 'k-', linewidth=1)
    
    # ====== GRÁFICO DE PIZZA (ESQUERDA) ======
    ax_pizza = fig.add_axes([0.1, 0.4, 0.3, 0.25])  # Reduzido de 0.35x0.3 para 0.3x0.25
    
    labels = list(dados_status.keys())
    valores = list(dados_status.values())
    
    cores = {
        'delivered': '#28a745',
        'undelivered': '#ffc107',
        'failed': '#dc3545',
        'sent': '#17a2b8',
        'queued': '#6c757d'
    }
    
    cores_grafico = [cores.get(label, '#95a5a6') for label in labels]
    
    # Gráfico de pizza compacto
    wedges, texts, autotexts = ax_pizza.pie(valores, labels=labels, autopct='%1.1f%%',
                                           colors=cores_grafico, startangle=90)
    
    # Melhorar aparência dos textos
    for text in texts:
        text.set_fontsize(8)  # Reduzido de 10 para 8
    for autotext in autotexts:
        autotext.set_fontsize(7)  # Reduzido de 9 para 7
        autotext.set_fontweight('bold')
        autotext.set_color('white')
    
    ax_pizza.set_title('Distribuição por Status', fontsize=10, fontweight='bold', pad=5)  # Reduzido de 12 para 10
    
    # ====== GRÁFICO DE BARRAS (DIREITA) ======
    ax_barras = fig.add_axes([0.55, 0.4, 0.3, 0.25])  # Reduzido de 0.35x0.3 para 0.3x0.25
    
    y_pos_bar = range(len(labels))
    bars = ax_barras.barh(y_pos_bar, valores, color=cores_grafico)
    
    # Configurar eixos
    ax_barras.set_yticks(y_pos_bar)
    ax_barras.set_yticklabels(labels)
    ax_barras.set_xlabel('Quantidade', fontsize=8)  # Reduzido de 10 para 8
    ax_barras.set_title('Quantidades Detalhadas', fontsize=10, fontweight='bold')  # Reduzido de 12 para 10
    
    # Adicionar valores nas barras
    for i, (bar, valor) in enumerate(zip(bars, valores)):
        ax_barras.text(bar.get_width() + max(valores) * 0.01, bar.get_y() + bar.get_height()/2,
                      f'{valor:,}', va='center', fontsize=7, fontweight='bold')  # Reduzido de 9 para 7
    
    # Ajustar layout das barras
    ax_barras.tick_params(axis='both', which='major', labelsize=7)  # Reduzido de 9 para 7
    
    # ====== RESUMO DOS STATUS (PARTE INFERIOR) ======
    ax_detalhes = fig.add_axes([0.1, 0.05, 0.8, 0.3])  # Aumentado altura de 0.2 para 0.3, movido para baixo
    ax_detalhes.set_xlim(0, 1)
    ax_detalhes.set_ylim(0, 1)
    ax_detalhes.axis('off')
    
    # Título da seção
    ax_detalhes.text(0.5, 0.9, 'DETALHAMENTO DOS STATUS', fontsize=14, fontweight='bold', 
                     ha='center', color='#2c3e50')
    
    # Lista dos status com descrições
    status_descricoes = {
        'delivered': 'Mensagens entregues com sucesso',
        'undelivered': 'Mensagens não entregues',
        'failed': 'Mensagens que falharam no envio',
        'sent': 'Mensagens enviadas (aguardando confirmação)',
        'queued': 'Mensagens na fila de envio'
    }
    
    y_start = 0.7
    for i, (status, quantidade) in enumerate(dados_status.items()):
        y_pos = y_start - (i * 0.08)  # Espaçamento ainda menor
        
        # Cor do status
        cor = cores.get(status, '#95a5a6')
        
        # Status, quantidade e porcentagem em uma linha
        porcentagem = (quantidade / total_filtrado * 100) if total_filtrado > 0 else 0
        descricao = status_descricoes.get(status, 'Status desconhecido')
        
        # Linha compacta com todos os dados
        ax_detalhes.text(0.05, y_pos, f'• {status.upper()}: {quantidade:,} mensagens ({porcentagem:.1f}%) - {descricao}', 
                        fontsize=10, fontweight='bold', color=cor)
    
    # Rodapé limpo (sem data/hora de geração)
    fig.text(0.5, 0.02, 'Campanha: Fortune Dragon · HiperBet',
             ha='center', fontsize=9, color='#6c757d')

    return fig

def gerar_relatorio_multiplo_pdf(dados_consolidados, resultados_por_arquivo):
    """Função principal para gerar o relatório PDF com múltiplos arquivos"""
    
    # Mostrar diretório atual
    diretorio_atual = os.getcwd()
    caminho_completo = os.path.join(diretorio_atual, ARQUIVO_PDF_SAIDA)
    
    print(f"📊 Gerando relatório PDF: {ARQUIVO_PDF_SAIDA}")
    print(f"📁 Diretório atual: {diretorio_atual}")
    print(f"📄 Caminho completo: {caminho_completo}")
    
    # Configurar estilo matplotlib para A4
    plt.style.use('default')
    plt.rcParams['font.size'] = 12
    plt.rcParams['axes.titlesize'] = 16
    # Evita interferência do tight_layout que estava causando sobreposição de títulos
    plt.rcParams['figure.autolayout'] = False
    
    try:
        with PdfPages(ARQUIVO_PDF_SAIDA) as pdf:
            # Página 1: Resumo consolidado
            print("📄 Criando página consolidada...")
            fig_consolidada = criar_pagina_consolidada(dados_consolidados, resultados_por_arquivo)
            # Mantém exatamente o tamanho A4 definido na figura
            pdf.savefig(fig_consolidada)
            plt.close(fig_consolidada)
            
        print(f"✅ Relatório PDF gerado: {ARQUIVO_PDF_SAIDA}")
        
        # Verificar se o arquivo foi criado
        if os.path.exists(ARQUIVO_PDF_SAIDA):
            tamanho = os.path.getsize(ARQUIVO_PDF_SAIDA)
            total_paginas = 1
            print(f"✅ Relatório gerado com sucesso: {ARQUIVO_PDF_SAIDA}")
            print(f"📂 Localização: {caminho_completo}")
            print(f"📏 Tamanho do arquivo: {tamanho:,} bytes")
            print(f"📄 O PDF contém {total_paginas} página com o resumo consolidado.")
        else:
            print(f"❌ ERRO: Arquivo não foi criado em {caminho_completo}")
        
    except Exception as e:
        print(f"❌ Erro ao gerar PDF: {e}")
        import traceback
        traceback.print_exc()


def gerar_pdf_relatorio(
    arquivos,
    *,
    filtro: Optional[str] = None,
    usar_regex: Optional[bool] = None,
    id_flow: Optional[str] = None,
    texto_sms: Optional[str] = None,
    arquivo_pdf_saida: Optional[str] = None,
    subtitle: Optional[str] = None,
    data_relatorio: Optional[str] = None,
):
    """
    Interface programática para gerar o PDF consolidado usando este módulo.
    Retorna o caminho absoluto do PDF gerado.
    """
    global FILTRO, USAR_REGEX, ID_FLOW, TEXTO_SMS, ARQUIVO_PDF_SAIDA, SUBTITLE, DATA_RELATORIO

    arquivos_list = [str(Path(arquivo)) for arquivo in arquivos if arquivo]
    if not arquivos_list:
        raise ValueError("Informe pelo menos um arquivo para gerar o relatório.")

    backup = {
        "FILTRO": FILTRO,
        "USAR_REGEX": USAR_REGEX,
        "ID_FLOW": ID_FLOW,
        "TEXTO_SMS": TEXTO_SMS,
        "ARQUIVO_PDF_SAIDA": ARQUIVO_PDF_SAIDA,
        "SUBTITLE": SUBTITLE,
        "DATA_RELATORIO": DATA_RELATORIO,
    }

    try:
        if filtro is not None:
            FILTRO = filtro
        if usar_regex is not None:
            USAR_REGEX = usar_regex
        if id_flow is not None:
            ID_FLOW = id_flow
        if arquivo_pdf_saida is not None:
            ARQUIVO_PDF_SAIDA = arquivo_pdf_saida
        if subtitle is not None:
            SUBTITLE = subtitle
        if data_relatorio is not None:
            DATA_RELATORIO = data_relatorio

        dados_consolidados, resultados_por_arquivo = analisar_multiplos_arquivos(arquivos_list)

        if not dados_consolidados["status_counter"]:
            raise ValueError("Nenhum dado encontrado com o filtro especificado. PDF não será gerado.")

        texto_dinamico = _resolver_texto_sms(texto_sms, FILTRO, resultados_por_arquivo)
        if texto_dinamico:
            TEXTO_SMS = texto_dinamico

        gerar_relatorio_multiplo_pdf(dados_consolidados, resultados_por_arquivo)
        return Path(ARQUIVO_PDF_SAIDA).expanduser().resolve()
    finally:
        FILTRO = backup["FILTRO"]
        USAR_REGEX = backup["USAR_REGEX"]
        ID_FLOW = backup["ID_FLOW"]
        TEXTO_SMS = backup["TEXTO_SMS"]
        ARQUIVO_PDF_SAIDA = backup["ARQUIVO_PDF_SAIDA"]
        SUBTITLE = backup["SUBTITLE"]
        DATA_RELATORIO = backup["DATA_RELATORIO"]

def main():
    print("🎯 ANALISADOR DE CAMPANHAS SMS - MÚLTIPLOS ARQUIVOS")
    print("=" * 60)
    
    # Descobrir arquivos para processar
    arquivos_para_processar = descobrir_arquivos(ARQUIVOS)
    
    if not arquivos_para_processar:
        print("❌ Nenhum arquivo encontrado para processar!")
        print(f"   Verifique a lista de arquivos: {ARQUIVOS}")
        return
    
    print(f"📁 Arquivos encontrados para processamento:")
    for arquivo in arquivos_para_processar:
        print(f"   • {arquivo}")
    print()
    
    # Analisar múltiplos arquivos
    dados_consolidados, resultados_por_arquivo = analisar_multiplos_arquivos(arquivos_para_processar)
    
    if not dados_consolidados['status_counter']:
        print("❌ Nenhum dado encontrado com o filtro especificado em nenhum arquivo. PDF não será gerado.")
        return
    
    # Gerar PDF
    print("\n" + "="*80)
    print("🎯 GERANDO RELATÓRIO CONSOLIDADO PDF...")
    print("="*80)
    
    gerar_relatorio_multiplo_pdf(dados_consolidados, resultados_por_arquivo)

if __name__ == "__main__":
    try:
        main()
    except ImportError as e:
        print(f"❌ Erro: Biblioteca não encontrada - {e}")
        print(f"💡 Instale as dependências com:")
        print(f"   pip install matplotlib pandas")
    except Exception as e:
        print(f"❌ Erro ao executar script: {e}")
        import traceback
        traceback.print_exc()
        print(f"💡 Instale as dependências com:")
        print(f"   pip install matplotlib pandas")
    except Exception as e:
        print(f"❌ Erro ao executar script: {e}")
        import traceback
        traceback.print_exc()
